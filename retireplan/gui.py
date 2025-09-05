from __future__ import annotations

import csv
import tkinter as tk
from pathlib import Path
from tkinter import ttk, filedialog, messagebox

from tksheet import Sheet
from ttkbootstrap import Style

from retireplan import inputs
from retireplan.engine import run_plan

# Column order for display
DISPLAY_COLUMNS = [
    "Year",
    "Your_Age",
    "Spouse_Age",
    "Lifestyle",
    "Filing",
    "Total_Spend",
    "Taxes Due",
    "Cash_Events",
    "Base_Spend",
    "Social_Security",
    "IRA_Draw",
    "Brokerage_Draw",
    "Roth_Draw",
    "Roth_Conversion",
    "RMD",
    "MAGI",
    "Sted_Deduction",
    "IRA_Balance",
    "Brokerage_Balance",
    "Roth Balance",
    "Total_Assets",
    "Shortfall",
]

# Hide these by default
HIDDEN_NAMES = {"MAGI", "Sted_Deduction", "Shortfall"}

# Engine → Display mapping
_DIRECT_MAP = {
    "Age_You": "Your_Age",
    "Age_Spouse": "Spouse_Age",
    "Phase": "Lifestyle",
    "Spend_Target": "Total_Spend",
    "Taxes": "Taxes Due",
    "Events_Cash": "Cash_Events",
    "Discretionary_Spend": "Base_Spend",
    "SS_Income": "Social_Security",
    "Draw_IRA": "IRA_Draw",
    "Draw_Brokerage": "Brokerage_Draw",
    "Draw_Roth": "Roth_Draw",
    "Roth_Conversion": "Roth_Conversion",
    "RMD": "RMD",
    "MAGI": "MAGI",
    "Std_Deduction": "Sted_Deduction",
    "End_Bal_IRA": "IRA_Balance",
    "End_Bal_Brokerage": "Brokerage_Balance",
    "End_Bal_Roth": "Roth Balance",
    "Total_Assets": "Total_Assets",
    "Shortfall": "Shortfall",
}


def _effective_filing(cfg, living_value: str) -> str:
    return (
        "MFJ" if (cfg.filing_status == "MFJ" and living_value == "Joint") else "Single"
    )


def _rows_to_display(cfg, rows: list[dict]) -> list[dict]:
    out: list[dict] = []
    for r in rows:
        d = {k: "" for k in DISPLAY_COLUMNS}
        d["Year"] = r.get("Year", "")
        d["Filing"] = _effective_filing(cfg, r.get("Living", "Single"))
        for src, dst in _DIRECT_MAP.items():
            if src in r:
                d[dst] = r[src]
        if not d.get("Lifestyle"):
            d["Lifestyle"] = r.get("Phase", "")
        out.append(d)
    return out


def _display_to_2d(rows_display: list[dict]) -> list[list]:
    return [[r.get(col, "") for col in DISPLAY_COLUMNS] for r in rows_display]


def _autosize_columns(sheet: Sheet, headers: list[str], data: list[list]) -> list[int]:
    """Set column widths by content. Return pixel widths per column."""
    widths: list[int] = []
    for c, h in enumerate(headers):
        maxlen = len(str(h))
        for row in data:
            s = str(row[c])
            if len(s) > maxlen:
                maxlen = len(s)
        px = max(80, min(260, 14 + maxlen * 7))  # heuristic
        widths.append(px)
        try:
            sheet.column_width(c, width=px)
        except Exception:
            pass
    return widths


def _apply_zebra(sheet: Sheet, nrows: int, *, mode: str):
    try:
        even_bg = "#f5f7fb" if mode == "Light" else "#1f2530"
        even = list(range(0, nrows, 2))
        sheet.dehighlight_all()
        sheet.highlight_rows(rows=even, bg=even_bg, fg=None)
    except Exception:
        pass


def _export_csv(path: Path, headers: list[str], data: list[list]):
    with path.open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow(headers)
        w.writerows(data)


def _export_xlsx(path: Path, headers: list[str], data: list[list]):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    wb = Workbook()
    ws = wb.active
    ws.title = "Projections"
    ws.append(headers)
    for row in data:
        ws.append(row)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="right")

    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        hdr = str(headers[col_idx - 1])
        maxlen = (
            max(len(hdr), *(len(str(r[col_idx - 1])) for r in data))
            if data
            else len(hdr)
        )
        ws.column_dimensions[col_letter].width = max(10, min(34, int(maxlen * 0.95)))

    last_row = ws.max_row
    last_col = ws.max_column
    tbl = Table(
        displayName="ProjectionsTable",
        ref=f"A1:{get_column_letter(last_col)}{last_row}",
    )
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(tbl)

    wb.save(path)


def _hide_initial_columns(sheet: Sheet, headers: list[str]) -> list[int]:
    """Hide default columns. Return hidden indices."""
    idx = [i for i, h in enumerate(headers) if h in HIDDEN_NAMES]
    if idx:
        try:
            sheet.hide_columns(idx)
        except Exception:
            pass
    return idx


def _fit_window_to_table(
    root: tk.Tk, sheet: Sheet, widths: list[int], hidden_idx: set[int], nrows: int
):
    """Resize sheet + window to show all columns and nrows rows."""
    visible_width = sum(w for i, w in enumerate(widths) if i not in hidden_idx)
    index_w = 52  # tksheet index column
    vscroll_w = 18
    hpad = 40  # borders/margins
    table_w = visible_width + index_w + vscroll_w + hpad

    # Row/height model
    row_h = 24
    header_h = 30
    top_controls_h = 56
    vpad = 40
    table_h = header_h + row_h * nrows + vpad

    # Screen clamp
    sw = root.winfo_screenwidth()
    sh = root.winfo_screenheight()
    win_w = min(table_w, sw - 40)
    win_h = min(top_controls_h + table_h, sh - 80)

    # Apply
    try:
        sheet.config(width=win_w - 16, height=win_h - top_controls_h - 16)
    except Exception:
        pass
    root.geometry(f"{int(win_w)}x{int(win_h)}")
    root.minsize(int(win_w), int(win_h))


class App:
    def __init__(self):
        self.style = Style("flatly")  # high-contrast light default
        self.root = self.style.master
        self.root.title("RetirePlan")

        self._build_ui()
        self._load_and_render()

    def _build_ui(self):
        top = ttk.Frame(self.root)
        top.pack(fill="x", padx=8, pady=6)

        ttk.Label(top, text="Filter:").pack(side="left", padx=(0, 6))
        self.filter_var = tk.StringVar(value="")
        self.filter_entry = ttk.Entry(top, textvariable=self.filter_var, width=32)
        self.filter_entry.pack(side="left")

        ttk.Label(top, text="Theme:").pack(side="left", padx=(14, 6))
        self.theme_var = tk.StringVar(value="Light")
        self.theme_box = ttk.Combobox(
            top,
            textvariable=self.theme_var,
            values=["Light", "Dark"],
            width=8,
            state="readonly",
        )
        self.theme_box.pack(side="left")

        self.btn_refresh = ttk.Button(top, text="Refresh", command=self._refresh)
        self.btn_csv = ttk.Button(top, text="Export CSV", command=self._export_csv)
        self.btn_xlsx = ttk.Button(top, text="Export XLSX", command=self._export_xlsx)
        self.btn_autosize = ttk.Button(top, text="Autosize", command=self._autosize_now)
        self.btn_quit = ttk.Button(top, text="Quit", command=self.root.destroy)
        for b in (
            self.btn_refresh,
            self.btn_csv,
            self.btn_xlsx,
            self.btn_autosize,
            self.btn_quit,
        ):
            b.pack(side="left", padx=6)

        frame = ttk.Frame(self.root)
        frame.pack(fill="both", expand=True, padx=8, pady=(0, 8))
        frame.grid_columnconfigure(0, weight=1)
        frame.grid_rowconfigure(0, weight=1)

        self.sheet = Sheet(
            frame,
            data=[],
            headers=DISPLAY_COLUMNS,
            height=540,
            width=1180,
            show_x_scrollbar=True,
            show_y_scrollbar=True,
        )
        self.sheet.grid(row=0, column=0, sticky="nsew")

        self.sheet.enable_bindings(
            (
                "single_select",
                "row_select",
                "column_select",
                "drag_select",
                "select_all",
                "column_width_resize",
                "double_click_column_resize",
                "row_height_resize",
                "arrowkeys",
                "right_click_popup_menu",
                "copy",
                "rc_select",
                "hide_columns",
            )
        )

        try:
            self.sheet.align(align="e")
        except Exception:
            pass

        self.filter_var.trace_add(
            "write", lambda *_: self._rebuild(self.filter_var.get())
        )
        self.theme_box.bind("<<ComboboxSelected>>", self._on_theme_change)

    def _apply_palette(self, mode: str):
        self.style.theme_use("flatly" if mode == "Light" else "darkly")
        opts_light = dict(
            table_bg="#ffffff",
            table_fg="#212529",
            table_grid_fg="#cfd4da",
            table_selected_cells_bg="#0d6efd",
            table_selected_cells_fg="#ffffff",
            header_bg="#e9ecef",
            header_fg="#212529",
            index_bg="#f8f9fa",
            index_fg="#212529",
            top_left_bg="#e9ecef",
            frame_bg="#ffffff",
        )
        opts_dark = dict(
            table_bg="#0f141b",
            table_fg="#e6edf3",
            table_grid_fg="#2b3441",
            table_selected_cells_bg="#0d6efd",
            table_selected_cells_fg="#ffffff",
            header_bg="#1b2430",
            header_fg="#e6edf3",
            index_bg="#151b23",
            index_fg="#e6edf3",
            top_left_bg="#1b2430",
            frame_bg="#0f141b",
        )
        opts = opts_light if mode == "Light" else opts_dark
        try:
            self.sheet.set_options(**opts)
        except Exception:
            pass
        self._apply_zebra(mode)

    def _apply_zebra(self, mode: str):
        data = self.sheet.get_sheet_data()
        _apply_zebra(self.sheet, len(data), mode=mode)

    def _load_and_render(self):
        cfg = inputs.load_yaml("examples/sample_inputs.yaml")
        rows = run_plan(cfg)
        display_rows = _rows_to_display(cfg, rows)
        data = _display_to_2d(display_rows)
        self.sheet.set_sheet_data(data, redraw=True)

        widths = _autosize_columns(self.sheet, DISPLAY_COLUMNS, data)
        hidden_idx = set(_hide_initial_columns(self.sheet, DISPLAY_COLUMNS))
        self._apply_palette(self.theme_var.get())

        # Fit window to all columns and all rows (cap at 30 rows shown)
        nrows = max(0, min(30, len(data)))
        _fit_window_to_table(self.root, self.sheet, widths, hidden_idx, nrows)

    def _rebuild(self, filter_text: str | None):
        cfg = inputs.load_yaml("examples/sample_inputs.yaml")
        rows = run_plan(cfg)
        display = _rows_to_display(cfg, rows)

        q = (filter_text or "").strip().lower()
        if q:
            display = [
                rd
                for rd in display
                if q in " ".join(str(rd.get(k, "")) for k in DISPLAY_COLUMNS).lower()
            ]

        data2 = _display_to_2d(display)
        self.sheet.set_sheet_data(data2, redraw=True)

        widths = _autosize_columns(self.sheet, DISPLAY_COLUMNS, data2)
        hidden_idx = set(_hide_initial_columns(self.sheet, DISPLAY_COLUMNS))
        self._apply_zebra(self.theme_var.get())

        nrows = max(0, min(30, len(data2)))
        _fit_window_to_table(self.root, self.sheet, widths, hidden_idx, nrows)

    def _refresh(self):
        self.filter_var.set("")
        self._load_and_render()

    def _autosize_now(self):
        data = self.sheet.get_sheet_data()
        widths = _autosize_columns(self.sheet, DISPLAY_COLUMNS, data)
        # Keep existing hidden set
        hidden_idx = {i for i, h in enumerate(DISPLAY_COLUMNS) if h in HIDDEN_NAMES}
        nrows = max(0, min(30, len(data)))
        _fit_window_to_table(self.root, self.sheet, widths, hidden_idx, nrows)

    def _export_csv(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv")],
            initialfile="projections.csv",
        )
        if not path:
            return
        try:
            current = self.sheet.get_sheet_data()
            _export_csv(Path(path), DISPLAY_COLUMNS, current)
            messagebox.showinfo("Export", f"Wrote {path}")
        except Exception as e:
            messagebox.showerror("Export failed", str(e))

    def _export_xlsx(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile="projections.xlsx",
        )
        if not path:
            return
        try:
            current = self.sheet.get_sheet_data()
            _export_xlsx(Path(path), DISPLAY_COLUMNS, current)
            messagebox.showinfo("Export", f"Wrote {path}")
        except Exception as e:
            messagebox.showerror("Export failed", str(e))

    def _on_theme_change(self, _event):
        self._apply_palette(self.theme_var.get())


def main() -> None:
    App().root.mainloop()


if __name__ == "__main__":
    main()
